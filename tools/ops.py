import tensorflow as tf
import os
import tensorflow.contrib.slim as slim
from openpyxl import workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import time
from PIL import Image
import cv2

'''创建文件'''
def create_file(summary_path):
    if not os.path.exists(summary_path):
        # print(summary_path)
        os.makedirs(summary_path)
        print('已创建文件：{}'.format(summary_path))

'''格式转换'''
def convert_type(ref,target):
    '''转格式'''
    target2 = tf.identity(target)
    ref2 = tf.identity(ref)

    converted_targets = tf.image.convert_image_dtype(target2, dtype=tf.uint8, saturate=True)
    converted_ref = tf.image.convert_image_dtype(ref2, dtype=tf.uint8, saturate=True)
    return converted_ref,converted_targets

'''计算psnr'''
def compute_psnr(ref, target):
    '''
    求psnr
    :param ref:参考图片，即原图
    :param target: 目标图，即网络输出图片
    :return:
    '''
    with tf.name_scope('compute_psnr'):
        '''转格式'''
        target2 = tf.identity(target)
        ref2 = tf.identity(ref)

        converted_targets = tf.image.convert_image_dtype(target2, dtype=tf.uint8, saturate=True)
        converted_ref = tf.image.convert_image_dtype(ref2, dtype=tf.uint8, saturate=True)

        '''计算PSNR'''
        ref3 = tf.cast(converted_ref, tf.float32)
        target3 = tf.cast(converted_targets, tf.float32)
        diff = target3 - ref3
        sqr = tf.multiply(diff, diff)
        err = tf.reduce_sum(sqr)
        v = tf.shape(diff)[0] * tf.shape(diff)[1] * tf.shape(diff)[2] * tf.shape(diff)[3]
        mse = err / tf.cast(v, tf.float32)
        psnr = 10. * (tf.log(255. * 255. / mse) / tf.log(10.))

        return psnr

'''prelu激活函数'''
def prelu_tf(inputs, name='Prelu'):
    '''prelu激活函数'''
    with tf.name_scope(name):
        with tf.variable_scope(name, reuse=tf.AUTO_REUSE):
            alphas = tf.get_variable('alpha', inputs.get_shape()[-1], initializer=tf.zeros_initializer(),
                                     dtype=tf.float32)
        pos = tf.nn.relu(inputs)
        neg = alphas * (inputs - abs(inputs)) * 0.5
    return pos + neg

'''添加可视化图像日志'''
def visual_image(self, name, input):
    '''添加可视化图片日志'''
    input_image = tf.reshape(tensor=input, shape=[-1, self.image_size, self.image_size, 3])
    tf.summary.image(name=name, tensor=input_image, max_outputs=4)

'''创建excel表格'''
def create_excel(excel_name):
    if os.path.exists(excel_name):
        excel = load_workbook(excel_name)
        excel_active = excel.active
    else:
        excel = workbook.Workbook()
        excel_active = excel.active
        excel_active.append(
            ['global_steps', 'epoch', 'step', 'loss_vae','psnr_transpose', 'psnr_skip', 'ssim_transpose', 'ssim_skip','Time','lr'])
    print('创建存储实验数据excel:{}'.format(excel_name))
    return excel,excel_active

'''存储psnr，ssir,loss_value数据到excel表中'''
def data_output(excel_active, global_steps, step, loss_value, psnr_first, psnr_second, ssim_first, ssim_second,Time,lr=0,epoch=0, ep=0):
    '''
    添加实验数据到excel表
    :param excel_active:
    :param global_steps:总的训练次数
    :param step:当前训练次数
    :param loss_value:vae总的损失
    :param psnr_first:第一阶段的psnr
    :param psnr_second:第二阶段的psnr
    :param ssim_first:第一阶段的psnr
    :param ssim_second:第二阶段的psnr
    :param epoch:迭代总数
    :param ep:当前迭代数
    :return:
    '''
    excel_active.append([global_steps, ep, step, loss_value, psnr_first, psnr_second, ssim_first, ssim_second,Time,lr])

    print('=' * 50)
    print('global_steps:{}, epoch:{}/{}, steps:{},'.format(global_steps, epoch,ep, step))
    print('-' * 50)
    print('current time:\t{}'.format(time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())))
    print('run time:\t\t%.3f' % Time)
    print('lr:\t\t\t\t{}'.format(lr))
    print('loss:\t\t\t{}'.format(loss_value))
    print('psnr_transpose:\t{}'.format(psnr_first))
    print('psnr_skip:\t\t{}'.format(psnr_second))
    print('ssim_transpose:\t{}'.format(ssim_first))
    print('ssim_skip:\t\t{}'.format(ssim_second))

'''存储训练结果到本地'''
def img_save_for_all(out_path, global_steps, image_size, image_org, img_an, first_out, second_out=None, is_train=True):
    '''
    图像保存到本地文件
    :param out_path:输出路径
    :param global_steps:总的训练次数
    :param image_size:
    :param image_org:
    :param img_an:
    :param first_out:第一阶段输出
    :param second_out:第二阶段输出
    :param is_train:判断是否为训练网络的存储
    :return:
    '''
    if is_train:
        for i in range(len(first_out)):
            plt.imsave(out_path + "/" + str(global_steps) + '_skip_' + str(i) + ".png",
                       np.reshape(second_out[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/"  + str(global_steps)+ '_org_' + str(i) + ".png",
                       np.reshape(image_org[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/"   + str(global_steps) + '_an_'+ str(i)+ ".png",
                       np.reshape(img_an[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) + '_transpose_' + str(i) + ".png",
                       np.reshape(first_out[i], newshape=[image_size, image_size, 3]))
    else:
        for i in range(len(second_out)):
            plt.imsave(out_path + "/" + str(global_steps) +'_' + str(i) + '_skip_' + ".png",
                       np.reshape(second_out[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) +'_'+ str(i) + '_org_' + ".png",
                       np.reshape(image_org[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) +'_'+ str(i) + '_an_' + ".png",
                       np.reshape(img_an[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) +'_' + str(i) + '_transpose_' + ".png",
                       np.reshape(first_out[i], newshape=[image_size, image_size, 3]))

def img_save_for_GR(out_path, global_steps, image_size, image_org, img_an, first_out, is_train=True):
    '''
    图像保存到本地文件，至存储第一阶段
    :param out_path:输出路径
    :param global_steps:总的训练次数
    :param image_size:
    :param image_org:
    :param img_an:
    :param first_out:第一阶段输出
    :param second_out:第二阶段输出
    :param is_train:判断是否为训练网络的存储
    :return:
    '''
    if is_train:
        for i in range(len(first_out)):
            # plt.imsave(out_path + "/" + str(global_steps) + '_skip_' + str(i) + ".png",
            #            np.reshape(second_out[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/"  + str(global_steps)+ '_org_' + str(i) + ".png",
                       np.reshape(image_org[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/"   + str(global_steps) + '_an_'+ str(i)+ ".png",
                       np.reshape(img_an[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) + '_transpose_' + str(i) + ".png",
                       np.reshape(first_out[i], newshape=[image_size, image_size, 3]))
    else:
        for i in range(len(first_out)):
            # plt.imsave(out_path + "/" + str(global_steps) +'_' + str(i) + '_skip_' + ".png",
            #            np.reshape(second_out[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) +'_'+ str(i) + '_org_' + ".png",
                       np.reshape(image_org[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) +'_'+ str(i) + '_an_' + ".png",
                       np.reshape(img_an[i], newshape=[image_size, image_size, 3]))
            plt.imsave(out_path + "/" + str(global_steps) +'_' + str(i) + '_transpose_' + ".png",
                       np.reshape(first_out[i], newshape=[image_size, image_size, 3]))

def fm_img_save(fm_names, fms_RGB, outpath, global_step):
    for i,fm_name in enumerate(fm_names):
        outpath2 = os.path.join(outpath,'feature_RGB',str(global_step))
        create_file(outpath2)
        outpath3 = outpath2 + '\\' + fm_name + '.png'

        t = fms_RGB[i]
        # s = fms_RGB[i].shape()
        # print(s)
        cv2.imwrite(outpath3,fms_RGB[i])
        plt.imsave(outpath3,fms_RGB[i])
        print('''{}'s feature map RGB image has save.'''.format(fm_name))

'''图像重命名'''
def img_rename(img_path):
    exposure = [-4,-3.5,-3,-2.5,-2,-1.5,-1,-0.5,0,0.5,1,1.5,2,2.5,3]
    no = 1
    index = 1
    imgs = os.listdir(img_path)
    for img in imgs:
        if no is 16:
            no = 1
            index += 1
        img_name = str(index) + '_' + str(exposure[no-1]) + '_512_0825.jpg'
        print(no)
        os.rename(os.path.join(img_path,img),os.path.join(img_path,img_name))

        if no == 9:
            img_fullpath = os.path.join(img_path,img_name)
            img = Image.open(img_fullpath)
            img.save('E:\WuXu\Dawn\MRVAE-Modification\data\img\\' + img_name)
            print('normal of {} has saved.'.format(img))
        no += 1

        print('{} has renamed.'.format(img))

def img_resize(img_path,img_w,img_h):
    for img_name in os.listdir(img_path):
        img_fullpath = os.path.join(img_path,img_name)
        img = Image.open(img_fullpath)
        img = img.resize((img_w,img_h))
        img.save(img_fullpath)
        print(img_name + ' has resized.')

# 转移图片
def img_move(inpath,outpath):
    import shutil
    create_file(outpath)

    list = os.listdir(inpath)

    for l in list:
        fp = os.path.join(inpath, l)
        list2 = os.listdir(fp)
        for l2 in list2:
            fp2 = os.path.join(fp, l2)
            op2 = os.path.join(outpath, l2)
            shutil.move(fp2, op2)


# 功能：将一字典写入到csv文件中
# 输入：文件名称，数据字典
def createDictCSV(fileName="", datalist=[]):
    with open(fileName, "w",newline='') as csvFile:
        csvWriter = csv.writer(csvFile)
        for k in range(len(datalist)):
            csvWriter.writerow([datalist[k][0],datalist[k][1],datalist[k][2],datalist[k][3],datalist[k][4]])
        csvFile.close()

'''MobileNet Conv'''
def depthwise_separable_conv(name,inputs,num_pwc_filters,width_multiplier,downsamle=False):
    '''
    depthwise_separable_conv
    :param name:
    :param inputs:
    :param num_pwc_filters:pointwise convolution的卷积核数量
    :param width_multiplier:控制输入的通道数
    :param downsamle:是否下采样
    :return:
    '''
    num_pwc_filters = round(num_pwc_filters * width_multiplier)     # 返回浮点数的四舍五入值
    stride = 2 if downsamle else 1
    depthwise_conv = slim.separable_convolution2d(inputs,num_outputs=None,stride=stride,depth_multiplier=1,kernel_size=[3,3])
    bn = slim.batch_norm(inputs=depthwise_conv)
    pointwise_conv = slim.convolution2d(bn,num_outputs=num_pwc_filters,kernel_size=[1,1])
    bn = slim.batch_norm(inputs=pointwise_conv)
    return bn

'''计算参数总量'''
def show_parament_numbers():
    from functools import reduce
    from operator import mul
    def get_num_params():
        num_params = 0
        for variable in tf.trainable_variables():
            shape = variable.get_shape()
            num_params += reduce(mul, [dim.value for dim in shape], 1)
        return num_params
    print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxx parament numbers is : %d' % get_num_params())

'''计算FLOPs、参数总量'''
# 源代码：https://robertlexis.github.io/2018/08/28/Tensorflow-%E6%A8%A1%E5%9E%8B%E6%B5%AE%E7%82%B9%E6%95%B0%E8%AE%A1%E7%AE%97%E9%87%8F%E5%92%8C%E5%8F%82%E6%95%B0%E9%87%8F%E7%BB%9F%E8%AE%A1/
def load_pb(pb):
    with tf.gfile.GFile(pb, "rb") as f:
        graph_def = tf.GraphDef()
        graph_def.ParseFromString(f.read())
    with tf.Graph().as_default() as graph:
        tf.import_graph_def(graph_def, name='')
        return graph

def stats_graph(graph):
    flops = tf.profiler.profile(graph, options=tf.profiler.ProfileOptionBuilder.float_operation())
    params = tf.profiler.profile(graph, options=tf.profiler.ProfileOptionBuilder.trainable_variables_parameter())
    print('FLOPs: {};    Trainable params: {}'.format(flops.total_float_ops, params.total_parameters))

'''图像改格式'''
def img_retype(img_path):

    imgs = os.listdir(img_path)
    for img in imgs:
        img_name = img[:-5] + '.jpg'
        os.rename(os.path.join(img_path,img),os.path.join(r'E:\WuXu\Dawn\MRVAE-Modification\tools\tp4',img_name))
        print('{} has retype.'.format(img))

if __name__=='__main__':
    img_retype(r'E:\WuXu\Dawn\MRVAE-Modification\tools\tp3')
    # img_rename(r'../data/eos_0825')
